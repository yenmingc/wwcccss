import httplib, urllib, base64, json
from openpyxl import Workbook
from openpyxl import load_workbook
from azure.storage.table import TableService, Entity

# This program reads controls from Excel and stores into an Azure Table
# Name of the file with cloud security standard
cfname = 'NIST 800-53-controls.xlsx'
# File with storage account info in json format
keyfile = 'Table.json'
with open(keyfile,'r') as k:
    t = json.load(k)
    acc = t['acc']
    key = t['key']
    tab = t['tab']

table_service = TableService(account_name=acc, account_key=key)

wb = load_workbook(filename = cfname)
ws = wb['Sheet1']

PK = 'US_NIST_800-53-r4'
row = 1
while not(ws['B'+str(row)].value is None and ws['F'+str(row)].value is None):
    row = row + 1
    rk = ws['B'+str(row)].value
    dm = ws['A'+str(row)].value
#    sdm = ws['B'+str(row)].value
    title = ws['C'+str(row)].value
    desc = ws['F'+str(row)].value
    sup = ws['G'+str(row)].value
    rel = ws['H'+str(row)].value
    control = {'PartitionKey': PK, 'RowKey': rk, 'Domain': dm, 'SupplementalGuide': sup, 'Title': title, 'ControlDescription': desc, 'Related': rel}
    table_service.insert_or_replace_entity(tab,control)
    print row, " ", control
